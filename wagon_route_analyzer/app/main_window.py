"""Main window for the wagon route report application."""

import os
import sys
from datetime import datetime

import pandas as pd
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QProgressBar, QDateEdit, QTableWidget, QTableWidgetItem,
    QMessageBox, QGroupBox, QGridLayout, QSplitter, QFrame,
    QHeaderView, QTextEdit, QTabWidget, QComboBox,
)

from .styles import STYLE
from .utils import clean_for_excel, safe_filename, fmt_dt
from .widgets import MultiSelectButton
from .workers import InitWorker, RouteReportWorker


COMBO_STYLE = (
    "QComboBox { background:#1a2035; color:#e2e8f0; border:1px solid #2d3748;"
    " border-radius:6px; padding:6px 10px; }"
    "QComboBox QAbstractItemView { background:#1a2035; color:#e2e8f0;"
    " selection-background-color:#2b4a7a; }"
)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.result_df = None
        self.worker = None
        self.setWindowTitle('Wagon Route Analyzer')
        self.setMinimumSize(1360, 860)
        self.setStyleSheet(STYLE)
        self._build_ui()
        self._load_reference_lists()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)

        hdr = QHBoxLayout()
        icon_lbl = QLabel('🛤️')
        icon_lbl.setFont(QFont('Segoe UI Emoji', 28))
        title_col = QVBoxLayout()
        title_lbl = QLabel('Wagon Route Analyzer')
        title_lbl.setObjectName('titleLabel')
        sub_lbl = QLabel('Route between stations · dwell time · distances · segment speed')
        sub_lbl.setObjectName('subtitleLabel')
        title_col.addWidget(title_lbl)
        title_col.addWidget(sub_lbl)
        hdr.addWidget(icon_lbl)
        hdr.addSpacing(10)
        hdr.addLayout(title_col)
        hdr.addStretch()
        root.addLayout(hdr)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet('color: #2d3748;')
        root.addWidget(line)

        settings_row = QHBoxLayout()
        settings_row.setSpacing(14)

        conn_box = QGroupBox('🔌 Database Connection')
        conn_grid = QGridLayout(conn_box)
        conn_grid.setSpacing(8)
        self.f_server   = self._field('your_server,port')
        self.f_database = self._field('your_database')
        self.f_username = self._field('your_username')
        self.f_password = self._field('your_password', password=True)
        for row, (lbl, fld) in enumerate([
            ('Server',   self.f_server),
            ('Database', self.f_database),
            ('Login',    self.f_username),
            ('Password', self.f_password),
        ]):
            conn_grid.addWidget(QLabel(lbl), row, 0)
            conn_grid.addWidget(fld, row, 1)
        settings_row.addWidget(conn_box, 3)

        period_box = QGroupBox('📅 Period')
        period_grid = QGridLayout(period_box)
        self.f_date_from = QDateEdit(QDate(QDate.currentDate().year(), 1, 1))
        self.f_date_from.setCalendarPopup(True)
        self.f_date_from.setDisplayFormat('dd.MM.yyyy')
        self.f_date_to = QDateEdit(QDate.currentDate())
        self.f_date_to.setCalendarPopup(True)
        self.f_date_to.setDisplayFormat('dd.MM.yyyy')
        period_grid.addWidget(QLabel('Date from'), 0, 0)
        period_grid.addWidget(self.f_date_from, 0, 1)
        period_grid.addWidget(QLabel('Date to'), 1, 0)
        period_grid.addWidget(self.f_date_to, 1, 1)
        settings_row.addWidget(period_box, 2)

        route_box = QGroupBox('🚉 Route & Filters')
        route_grid = QGridLayout(route_box)
        route_grid.setSpacing(8)

        self.f_load_station = QComboBox()
        self.f_load_station.setEditable(True)
        self.f_load_station.addItem('All')
        self.f_load_station.setMinimumWidth(220)
        self.f_load_station.setStyleSheet(COMBO_STYLE)

        self.f_unload_station = QComboBox()
        self.f_unload_station.setEditable(True)
        self.f_unload_station.addItem('All')
        self.f_unload_station.setMinimumWidth(220)
        self.f_unload_station.setStyleSheet(COMBO_STYLE)

        self.f_manager = QComboBox()
        self.f_manager.addItem('All')
        self.f_manager.setMinimumWidth(220)
        self.f_manager.setStyleSheet(COMBO_STYLE)

        self.f_type = MultiSelectButton('All types')
        self.f_type.setMinimumWidth(220)

        route_grid.addWidget(QLabel('Loading station'),   0, 0)
        route_grid.addWidget(self.f_load_station,         0, 1)
        route_grid.addWidget(QLabel('Unloading station'), 1, 0)
        route_grid.addWidget(self.f_unload_station,       1, 1)
        route_grid.addWidget(QLabel('Manager'),           2, 0)
        route_grid.addWidget(self.f_manager,              2, 1)
        route_grid.addWidget(QLabel('Wagon type'),        3, 0)
        route_grid.addWidget(self.f_type,                 3, 1)
        settings_row.addWidget(route_box, 3)

        btn_box = QGroupBox('⚡ Actions')
        btn_col = QVBoxLayout(btn_box)
        self.run_btn = QPushButton('🚀  Build route')
        self.run_btn.setObjectName('runBtn')
        self.run_btn.setMinimumHeight(48)
        self.run_btn.clicked.connect(self.start_calculation)
        self.export_btn = QPushButton('💾  Export to Excel')
        self.export_btn.setObjectName('exportBtn')
        self.export_btn.setMinimumHeight(40)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_excel)
        self.refresh_btn = QPushButton('↻  Refresh lists')
        self.refresh_btn.setObjectName('exportBtn')
        self.refresh_btn.clicked.connect(self._load_reference_lists)
        btn_col.addWidget(self.run_btn)
        btn_col.addWidget(self.export_btn)
        btn_col.addWidget(self.refresh_btn)
        btn_col.addStretch()
        settings_row.addWidget(btn_box, 2)
        root.addLayout(settings_row)

        prog_row = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.status_lbl = QLabel('Ready')
        self.status_lbl.setStyleSheet('color: #718096; font-size: 12px;')
        prog_row.addWidget(self.progress_bar, 4)
        prog_row.addWidget(self.status_lbl, 1)
        root.addLayout(prog_row)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(10)
        self.stat_wagons     = self._stat_card('Wagons', '—')
        self.stat_routes     = self._stat_card('Routes', '—')
        self.stat_rows       = self._stat_card('Station rows', '—')
        self.stat_long_stops = self._stat_card('Stops > 24 h', '—')
        for w in [self.stat_wagons, self.stat_routes, self.stat_rows, self.stat_long_stops]:
            stats_row.addWidget(w)
        root.addLayout(stats_row)

        splitter = QSplitter(Qt.Vertical)
        tabs = QTabWidget()
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            self.table.styleSheet() + 'QTableWidget { alternate-background-color: #1a2035; }')
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        tabs.addTab(self.table, '📋 Route data')

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(160)
        tabs.addTab(self.log_box, '🖥️ Log')

        splitter.addWidget(tabs)
        root.addWidget(splitter)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _field(self, placeholder='', password=False):
        f = QLineEdit()
        f.setPlaceholderText(placeholder)
        if password:
            f.setEchoMode(QLineEdit.Password)
        return f

    def _stat_card(self, title, value):
        lbl = QLabel(f'<b style="font-size:18px">{value}</b><br>'
                     f'<span style="font-size:11px;color:#718096">{title}</span>')
        lbl.setObjectName('statLabel')
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setMinimumWidth(130)
        return lbl

    def _update_stat(self, lbl, title, value):
        lbl.setText(f'<b style="font-size:18px">{value}</b><br>'
                    f'<span style="font-size:11px;color:#718096">{title}</span>')

    def _conn_str(self):
        s = self.f_server.text().strip()
        d = self.f_database.text().strip()
        u = self.f_username.text().strip()
        p = self.f_password.text().strip()
        if u and p:
            return f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={s};DATABASE={d};UID={u};PWD={p}'
        return f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={s};DATABASE={d};Trusted_Connection=yes'

    # ------------------------------------------------------------------
    # Reference data loading
    # ------------------------------------------------------------------

    def _load_reference_lists(self):
        self.status_lbl.setText('Loading reference data...')
        self.run_btn.setEnabled(False)
        self._init_worker = InitWorker(self._conn_str())
        self._init_worker.companies_loaded.connect(self._on_companies)
        self._init_worker.types_loaded.connect(self._on_types)
        self._init_worker.stations_loaded.connect(self._on_stations)
        self._init_worker.error.connect(self._on_init_error)
        self._init_worker.finished.connect(lambda: (
            self.run_btn.setEnabled(True),
            self.status_lbl.setText('Ready'),
        ))
        self._init_worker.start()

    def _on_companies(self, items):
        self.f_manager.clear()
        self.f_manager.addItem('All')
        self.f_manager.addItems(items)

    def _on_types(self, items):
        self.f_type.set_items(items)

    def _on_stations(self, items):
        current_load   = self.f_load_station.currentText()
        current_unload = self.f_unload_station.currentText()
        self.f_load_station.clear()
        self.f_load_station.addItem('All')
        self.f_load_station.addItems(items)
        self.f_unload_station.clear()
        self.f_unload_station.addItem('All')
        self.f_unload_station.addItems(items)
        if current_load:
            idx = self.f_load_station.findText(current_load)
            if idx >= 0:
                self.f_load_station.setCurrentIndex(idx)
        if current_unload:
            idx = self.f_unload_station.findText(current_unload)
            if idx >= 0:
                self.f_unload_station.setCurrentIndex(idx)

    def _on_init_error(self, msg):
        self.run_btn.setEnabled(True)
        self.status_lbl.setText('Reference data error')
        self._on_log(f'Failed to load reference data: {msg}')

    # ------------------------------------------------------------------
    # Calculation
    # ------------------------------------------------------------------

    def start_calculation(self):
        date_from = self.f_date_from.date().toString('yyyyMMdd')
        date_to   = self.f_date_to.date().toString('yyyyMMdd')
        if self.f_date_to.date() < self.f_date_from.date():
            QMessageBox.warning(self, 'Date error', "The 'to' date must be >= the 'from' date.")
            return

        self.run_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_box.clear()
        self.table.setRowCount(0)
        self.result_df = None

        self.worker = RouteReportWorker(
            self._conn_str(),
            date_from,
            date_to,
            load_station=self.f_load_station.currentText(),
            unload_station=self.f_unload_station.currentText(),
            manager=self.f_manager.currentText(),
            wagon_types=self.f_type.selected(),
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_finished)
        self.worker.error.connect(self._on_error)
        self.worker.log.connect(self._on_log)
        self.worker.start()

    def _on_progress(self, pct, msg):
        self.progress_bar.setValue(pct)
        self.status_lbl.setText(msg)

    def _on_log(self, msg):
        self.log_box.append(msg)

    def _on_error(self, msg):
        self.run_btn.setEnabled(True)
        self.status_lbl.setText('Error!')
        QMessageBox.critical(self, 'Error', f'An error occurred:\n\n{msg[:3000]}')
        self._on_log(f'ERROR:\n{msg}')

    def _on_finished(self, df):
        self.result_df = df
        self.run_btn.setEnabled(True)
        if df is None or df.empty:
            self.status_lbl.setText('No routes found')
            self._fill_table(pd.DataFrame())
            self._update_stats_from(pd.DataFrame())
            return
        self.export_btn.setEnabled(True)
        self.status_lbl.setText('Done')
        self._fill_table(df)
        self._update_stats_from(df)

    # ------------------------------------------------------------------
    # Stats & table
    # ------------------------------------------------------------------

    def _update_stats_from(self, df):
        if df is None or df.empty:
            for card, title in [
                (self.stat_wagons, 'Wagons'),
                (self.stat_routes, 'Routes'),
                (self.stat_rows,   'Station rows'),
                (self.stat_long_stops, 'Stops > 24 h'),
            ]:
                self._update_stat(card, title, '0')
            return
        wagons     = df['Wagon number'].nunique()     if 'Wagon number'     in df.columns else 0
        routes     = df['Route ID'].nunique()         if 'Route ID'         in df.columns else 0
        rows       = len(df)
        long_stops = int((df.get('Stop > 24 h', '') == 'Yes').sum()) if 'Stop > 24 h' in df.columns else 0
        self._update_stat(self.stat_wagons,     'Wagons',       f'{wagons:,}')
        self._update_stat(self.stat_routes,     'Routes',       f'{routes:,}')
        self._update_stat(self.stat_rows,       'Station rows', f'{rows:,}')
        self._update_stat(self.stat_long_stops, 'Stops > 24 h', f'{long_stops:,}')

    def _fill_table(self, df):
        if df is None or df.empty:
            self.table.setColumnCount(0)
            self.table.setRowCount(0)
            return

        display_df = df.copy()
        date_cols = [
            'Departure date',
            'First operation at station',
            'Last operation at station',
        ]
        for col in date_cols:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(fmt_dt)

        limit = min(len(display_df), 5000)
        self.table.setColumnCount(len(display_df.columns))
        self.table.setHorizontalHeaderLabels(display_df.columns.tolist())
        self.table.setRowCount(limit)

        yellow    = QColor(255, 242, 153)
        dark_text = QColor(25, 25, 25)
        for r in range(limit):
            is_long_stop = str(display_df.iloc[r].get('Stop > 24 h', '')) == 'Yes'
            for c, col in enumerate(display_df.columns):
                val  = display_df.iloc[r, c]
                item = QTableWidgetItem(
                    '' if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val))
                item.setTextAlignment(Qt.AlignCenter)
                if is_long_stop and col in {'Station', 'Dwell time, h', 'Stop > 24 h'}:
                    item.setBackground(yellow)
                    item.setForeground(dark_text)
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_excel(self):
        if self.result_df is None or self.result_df.empty:
            return
        try:
            base_dir = os.path.dirname(
                sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
            suffix_parts = []
            if self.f_load_station.currentText() not in ('All', ''):
                suffix_parts.append('load_' + safe_filename(self.f_load_station.currentText()[:25]))
            if self.f_unload_station.currentText() not in ('All', ''):
                suffix_parts.append('unload_' + safe_filename(self.f_unload_station.currentText()[:25]))
            suffix = '_' + '_'.join(suffix_parts) if suffix_parts else ''
            fname  = f"wagon_route{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path   = os.path.join(base_dir, fname)

            df_exp = self.result_df.copy()
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                clean_for_excel(df_exp).to_excel(writer, index=False, sheet_name='Route')
                ws = writer.sheets['Route']
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                thin   = Side(style='thin', color='2D3748')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                hdr_fill = PatternFill('solid', fgColor='1E3A5F')
                hdr_font = Font(bold=True, color='FFFFFF', size=11)
                alt_fill = PatternFill('solid', fgColor='F0F4F8')
                yellow   = PatternFill('solid', fgColor='FFF299')

                headers    = list(df_exp.columns)
                station_col = headers.index('Station') + 1        if 'Station'        in headers else None
                hours_col   = headers.index('Dwell time, h') + 1  if 'Dwell time, h'  in headers else None
                flag_col    = headers.index('Stop > 24 h') + 1    if 'Stop > 24 h'    in headers else None

                for ci, col in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=ci)
                    cell.fill      = hdr_fill
                    cell.font      = hdr_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border    = border

                for ri in range(2, len(df_exp) + 2):
                    is_long = str(ws.cell(row=ri, column=flag_col).value) == 'Yes' if flag_col else False
                    for ci in range(1, len(headers) + 1):
                        cell           = ws.cell(row=ri, column=ci)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border    = border
                        if is_long and ci in {station_col, hours_col, flag_col}:
                            cell.fill = yellow
                        elif ri % 2 == 0:
                            cell.fill = alt_fill

                for ci, col in enumerate(headers, 1):
                    values = [str(col)] + [str(v) for v in df_exp[col].head(200).fillna('').tolist()]
                    width  = min(max(len(v) for v in values) + 2, 45)
                    ws.column_dimensions[get_column_letter(ci)].width = width
                ws.freeze_panes       = 'A2'
                ws.row_dimensions[1].height = 42
                ws.auto_filter.ref    = ws.dimensions

            self._on_log(f'✅ File saved: {path}')
            msg      = QMessageBox(self)
            msg.setWindowTitle('Done!')
            msg.setText('File saved:')
            msg.setInformativeText(path)
            msg.setIcon(QMessageBox.Information)
            open_btn = msg.addButton('📂  Open folder', QMessageBox.ActionRole)
            msg.addButton('OK', QMessageBox.AcceptRole)
            msg.exec_()
            if msg.clickedButton() == open_btn:
                os.startfile(base_dir)
        except Exception as e:
            QMessageBox.critical(self, 'Save error', str(e))
            self._on_log(f'❌ Error: {e}')
